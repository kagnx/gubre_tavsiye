"""Excel ve CSV dosyalarindan veri yukleme."""
import csv
import os


def csv_yukle(dosya_yolu):
    """CSV dosyasindan toprak analiz verilerini yukler.
    Beklenen format: Bolge, Bitki, Tarim_Sekli, OM, P, K
    veya tek satir: OM, P, K
    """
    sonuclar = []
    try:
        with open(dosya_yolu, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for satir in reader:
                kayit = {}
                # Esnek sutun eslestirme
                for anahtar, deger in satir.items():
                    alt = anahtar.strip().lower()
                    if alt in ("bolge", "bölge", "region"):
                        kayit["bolge"] = deger.strip()
                    elif alt in ("bitki", "ürün", "crop", "plant"):
                        kayit["bitki"] = deger.strip()
                    elif alt in ("tarim_sekli", "tarım şekli", "farming", "technique"):
                        kayit["tarim_sekli"] = deger.strip()
                    elif alt in ("om", "organik_madde", "organik madde", "organic_matter", "organic matter"):
                        kayit["om"] = float(deger.strip().replace(",", "."))
                    elif alt in ("p", "fosfor", "phosphorus", "p2o5"):
                        kayit["fosfor"] = float(deger.strip().replace(",", "."))
                    elif alt in ("k", "potasyum", "potassium", "k2o"):
                        kayit["potasyum"] = float(deger.strip().replace(",", "."))
                if "om" in kayit and "fosfor" in kayit and "potasyum" in kayit:
                    sonuclar.append(kayit)
    except Exception as e:
        return None, str(e)
    return sonuclar, None


def excel_yukle(dosya_yolu):
    """Excel dosyasindan toprak analiz verilerini yukler."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(dosya_yolu, read_only=True, data_only=True)
        ws = wb.active

        # Baslik satirini bul
        basliklar = []
        for hucre in next(ws.iter_rows(min_row=1, max_row=1)):
            basliklar.append(str(hucre.value or "").strip().lower())

        sonuclar = []
        for satir in ws.iter_rows(min_row=2, values_only=True):
            kayit = {}
            for i, deger in enumerate(satir):
                if i < len(basliklar):
                    alt = basliklar[i]
                    if deger is None:
                        continue
                    deger_str = str(deger).strip()
                    if alt in ("bolge", "bölge", "region"):
                        kayit["bolge"] = deger_str
                    elif alt in ("bitki", "ürün", "crop", "plant"):
                        kayit["bitki"] = deger_str
                    elif alt in ("tarim_sekli", "tarım şekli", "farming", "technique"):
                        kayit["tarim_sekli"] = deger_str
                    elif alt in ("om", "organik_madde", "organik madde", "organic_matter"):
                        kayit["om"] = float(deger_str.replace(",", "."))
                    elif alt in ("p", "fosfor", "phosphorus", "p2o5"):
                        kayit["fosfor"] = float(deger_str.replace(",", "."))
                    elif alt in ("k", "potasyum", "potassium", "k2o"):
                        kayit["potasyum"] = float(deger_str.replace(",", "."))
            if "om" in kayit and "fosfor" in kayit and "potasyum" in kayit:
                sonuclar.append(kayit)

        wb.close()
        return sonuclar, None
    except Exception as e:
        return None, str(e)


def dosya_yukle(dosya_yolu):
    """Dosya uzantisina gore otomatik yukleme yapar."""
    ext = os.path.splitext(dosya_yolu)[1].lower()
    if ext == ".csv":
        return csv_yukle(dosya_yolu)
    elif ext in (".xlsx", ".xls"):
        return excel_yukle(dosya_yolu)
    else:
        return None, "Desteklenmeyen dosya formati"
